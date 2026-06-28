#!/usr/bin/env python3
"""Scan a directory of card images and write OCR results to an Excel spreadsheet."""

from __future__ import annotations  # Forward refs without quotes

import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from ocrmac import ocrmac

from . import __version__


# Column headers for the output sheet

G_STR_COL_FILE = "File"
G_STR_COL_TEXT = "Text"
G_STR_COL_CONFIDENCE = "Confidence"
G_STR_COL_BBOX = "Bounding Box"


def LAnnotationOcrImage(
	pathImage: Path,
	fLiveText: bool = True,
) -> list[tuple[str, float, list[float]]]:
	# Returns list of (text, confidence, bbox) for all recognized text in the image.
	# bbox is normalized [x, y, w, h] with origin at bottom-left (Vision convention).

	strFramework = "livetext" if fLiveText else "vision"

	return ocrmac.OCR(str(pathImage), framework=strFramework, unit="line").recognize()


def WriteOcrResultsXlsx(pathDir: Path, pathOutput: Path, fLiveText: bool = True) -> None:
	# Scans all PNG images in pathDir, runs OCR on each, and writes results to pathOutput.
	# Each annotation gets its own row: filename, recognized text, confidence, bbox.

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "OCR Results"

	# Header row

	lStrHeaders = [G_STR_COL_FILE, G_STR_COL_TEXT, G_STR_COL_CONFIDENCE, G_STR_COL_BBOX]
	ws.append(lStrHeaders)

	fontBold = Font(name="Arial", bold=True)
	for cell in ws[1]:
		cell.font = fontBold

	# Data rows — one row per annotation per image

	for pathImage in sorted(pathDir.glob("*.png")):
		lAnnotation = LAnnotationOcrImage(pathImage, fLiveText=fLiveText)

		for strText, gConf, lBbox in lAnnotation:
			strBbox = ", ".join(f"{g:.4f}" for g in lBbox)
			ws.append([pathImage.name, strText, round(gConf, 4), strBbox])

	# Column widths and alignment

	ws.column_dimensions["A"].width = 30
	ws.column_dimensions["B"].width = 50
	ws.column_dimensions["C"].width = 12
	ws.column_dimensions["D"].width = 36

	alignWrap = Alignment(wrap_text=True, vertical="top")
	for row in ws.iter_rows(min_row=2):
		for cell in row:
			cell.font = Font(name="Arial")
			cell.alignment = alignWrap

	wb.save(pathOutput)


def main() -> None:
	parser = argparse.ArgumentParser(
		prog="ocr",
		description="Get text from card images using Apple Vision OCR.",
	)
	parser.add_argument(
		"--version",
		action="version",
		version=f"%(prog)s {__version__}",
	)
	parser.add_argument(
		"--images-dir",
		type=Path,
		default=Path("images"),
		help="Directory containing PNG card images (default: 'images')",
	)
	parser.add_argument(
		"--output",
		type=Path,
		default=Path("ocr_results.xlsx"),
		help="Output Excel file path (default: 'ocr_results.xlsx')",
	)
	parser.add_argument(
		"--no-livetext",
		action="store_true",
		help="Use Vision framework instead of LiveText (slower, returns per-token confidence)",
	)

	args = parser.parse_args()

	pathDir: Path = args.images_dir
	pathOutput: Path = args.output
	fLiveText: bool = not args.no_livetext

	if not pathDir.is_dir():
		print(f"error: images directory not found: {pathDir}", file=sys.stderr)
		sys.exit(1)

	lPathImages = sorted(pathDir.glob("*.png"))

	if not lPathImages:
		print(f"error: no PNG images found in {pathDir}", file=sys.stderr)
		sys.exit(1)

	print(f"scanning {len(lPathImages)} images in '{pathDir}'...")
	WriteOcrResultsXlsx(pathDir, pathOutput, fLiveText=fLiveText)
	print(f"results written to '{pathOutput}'")


if __name__ == "__main__":
	main()